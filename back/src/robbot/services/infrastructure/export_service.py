"""
Export Service -  L1 (Refatorado .1)

Serviço para exportação de relatórios em PDF e Excel.
Error handling adicionado para debugging.
"""

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

logger = logging.getLogger(__name__)


class ExportService:
    """Serviço para exportação de relatórios em PDF/Excel"""

    # =============================================================================
    # PDF EXPORTS
    # =============================================================================

    @staticmethod
    def export_performance_report_pdf(report_data: dict[str, Any]) -> bytes:
        """
        Exporta relatório de performance em PDF.

        Args:
            report_data: Dict com estrutura do PerformanceReportSchema

        Returns:
            bytes: PDF em memória
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1a73e8"),
            spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#333333"),
            spaceAfter=8,
        )

        elements = []

        # Título
        period = report_data.get("period", {})
        title_text = "Relatório de Performance de Atendimento"
        elements.append(Paragraph(title_text, title_style))
        elements.append(
            Paragraph(
                f"Período: {period.get('start')} até {period.get('end')}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 0.3 * inch))

        # Seção 1: Tempo de Resposta do Bot
        bot_response = report_data.get("bot_response_time", {})
        elements.append(Paragraph("1. Tempo de Resposta do Bot (LLM)", heading_style))

        response_data = [
            ["Métrica", "Valor"],
            ["Média (ms)", f"{bot_response.get('avg_ms', 0):.2f}"],
            ["Mediana (ms)", f"{bot_response.get('median_ms', 0):.2f}"],
            ["P95 (ms)", f"{bot_response.get('p95_ms', 0):.2f}"],
            ["P99 (ms)", f"{bot_response.get('p99_ms', 0):.2f}"],
            ["Mínimo (ms)", f"{bot_response.get('min_ms', 0)}"],
            ["Máximo (ms)", f"{bot_response.get('max_ms', 0)}"],
            [
                "Total de Interações",
                f"{bot_response.get('total_interactions', 0):,}",
            ],
        ]

        response_table = Table(response_data, colWidths=[3 * inch, 2 * inch])
        response_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(response_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Seção 2: Taxa de Handoff
        handoff_stats = report_data.get("handoff_stats", {})
        elements.append(Paragraph("2. Taxa de Resolução e Handoff", heading_style))

        handoff_data = [
            ["Métrica", "Valor"],
            [
                "Total de Conversas",
                f"{handoff_stats.get('total_conversations', 0):,}",
            ],
            ["Resolvidas pelo Bot", f"{handoff_stats.get('bot_resolved', 0):,}"],
            ["Handoff para Humano", f"{handoff_stats.get('handoff_required', 0):,}"],
            [
                "Taxa de Resolução Automática",
                f"{handoff_stats.get('auto_resolution_rate', 0):.2f}%",
            ],
            ["Taxa de Handoff", f"{handoff_stats.get('handoff_rate', 0):.2f}%"],
        ]

        handoff_table = Table(handoff_data, colWidths=[3 * inch, 2 * inch])
        handoff_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(handoff_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Seção 3: Horários de Pico
        peak_hours = report_data.get("peak_hours", [])
        if peak_hours:
            elements.append(Paragraph("3. Horários de Pico de Atendimento", heading_style))

            peak_data = [["Hora", "Mensagens", "Conversas"]]
            for hour_data in peak_hours[:10]:  # Top 10 horas
                peak_data.append(
                    [
                        f"{hour_data.get('hour')}:00",
                        f"{hour_data.get('message_count', 0):,}",
                        f"{hour_data.get('conversation_count', 0):,}",
                    ]
                )

            peak_table = Table(peak_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch])
            peak_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            elements.append(peak_table)
            elements.append(PageBreak())

        # Seção 4: Distribuição por Status
        status_distribution = report_data.get("status_distribution", [])
        if status_distribution:
            elements.append(Paragraph("4. Distribuição de Conversas por Status", heading_style))

            status_data = [["Status", "Quantidade", "Percentual"]]
            for status in status_distribution:
                status_data.append(
                    [
                        status.get("status", "N/A"),
                        f"{status.get('count', 0):,}",
                        f"{status.get('percentage', 0):.2f}%",
                    ]
                )

            status_table = Table(status_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch])
            status_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            elements.append(status_table)

        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        footer_text = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}"
        elements.append(Paragraph(footer_text, styles["Normal"]))

        # Build PDF
        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        logger.info("[SUCCESS] PDF gerado com sucesso (%s bytes)", len(pdf_bytes))
        return pdf_bytes

    # =============================================================================
    # EXCEL EXPORTS
    # =============================================================================

    @staticmethod
    def export_performance_report_excel(report_data: dict[str, Any]) -> bytes:
        """
        Exporta relatório de performance em Excel.

        Args:
            report_data: Dict com estrutura do PerformanceReportSchema

        Returns:
            bytes: Excel em memória
        """
        wb = Workbook()

        # Estilos
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Sheet 1: Resumo
        ws_summary = wb.active
        ws_summary.title = "Resumo"  # type: ignore[assignment]

        period = report_data.get("period", {})
        ws_summary["A1"] = "Relatório de Performance de Atendimento"
        ws_summary["A1"].font = Font(bold=True, size=16)  # type: ignore[index]
        ws_summary["A2"] = f"Período: {period.get('start')} até {period.get('end')}"

        # Bot Response Time
        bot_response = report_data.get("bot_response_time", {})
        ws_summary["A4"] = "Tempo de Resposta do Bot (LLM)"
        ws_summary["A4"].font = Font(bold=True, size=14)  # type: ignore[index]

        headers = ["Métrica", "Valor"]
        for col, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=5, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        metrics_data = [
            ["Média (ms)", bot_response.get("avg_ms", 0)],
            ["Mediana (ms)", bot_response.get("median_ms", 0)],
            ["P95 (ms)", bot_response.get("p95_ms", 0)],
            ["P99 (ms)", bot_response.get("p99_ms", 0)],
            ["Mínimo (ms)", bot_response.get("min_ms", 0)],
            ["Máximo (ms)", bot_response.get("max_ms", 0)],
            ["Total de Interações", bot_response.get("total_interactions", 0)],
        ]

        for row_idx, (metric, value) in enumerate(metrics_data, start=6):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
            ws_summary.cell(row=row_idx, column=1).fill = data_fill  # type: ignore[index]
            ws_summary.cell(row=row_idx, column=2).fill = data_fill  # type: ignore[index]

        # Handoff Stats
        handoff_stats = report_data.get("handoff_stats", {})
        ws_summary["A14"] = "Taxa de Resolução e Handoff"
        ws_summary["A14"].font = Font(bold=True, size=14)  # type: ignore[index]

        for col, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=15, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        handoff_data = [
            ["Total de Conversas", handoff_stats.get("total_conversations", 0)],
            ["Resolvidas pelo Bot", handoff_stats.get("bot_resolved", 0)],
            ["Handoff para Humano", handoff_stats.get("handoff_required", 0)],
            [
                "Taxa de Resolução Automática (%)",
                handoff_stats.get("auto_resolution_rate", 0),
            ],
            ["Taxa de Handoff (%)", handoff_stats.get("handoff_rate", 0)],
        ]

        for row_idx, (metric, value) in enumerate(handoff_data, start=16):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
            ws_summary.cell(row=row_idx, column=1).fill = data_fill  # type: ignore[index]
            ws_summary.cell(row=row_idx, column=2).fill = data_fill  # type: ignore[index]

        # Ajustar largura das colunas
        ws_summary.column_dimensions["A"].width = 35  # type: ignore[index]
        ws_summary.column_dimensions["B"].width = 20  # type: ignore[index]

        # Sheet 2: Horários de Pico
        ws_peak = wb.create_sheet(title="Horários de Pico")
        peak_hours = report_data.get("peak_hours", [])

        ws_peak["A1"] = "Horários de Pico de Atendimento"
        ws_peak["A1"].font = Font(bold=True, size=14)  # type: ignore[index]

        peak_headers = ["Hora", "Total de Mensagens", "Total de Conversas"]
        for col, header in enumerate(peak_headers, start=1):
            cell = ws_peak.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, hour_data in enumerate(peak_hours, start=4):
            ws_peak.cell(row=row_idx, column=1, value=f"{hour_data.get('hour')}:00")
            ws_peak.cell(row=row_idx, column=2, value=hour_data.get("message_count", 0))
            ws_peak.cell(row=row_idx, column=3, value=hour_data.get("conversation_count", 0))
            ws_peak.cell(row=row_idx, column=1).fill = data_fill  # type: ignore[index]
            ws_peak.cell(row=row_idx, column=2).fill = data_fill  # type: ignore[index]
            ws_peak.cell(row=row_idx, column=3).fill = data_fill  # type: ignore[index]

        ws_peak.column_dimensions["A"].width = 15  # type: ignore[index]
        ws_peak.column_dimensions["B"].width = 20  # type: ignore[index]
        ws_peak.column_dimensions["C"].width = 20  # type: ignore[index]

        # Sheet 3: Status Distribution
        ws_status = wb.create_sheet(title="Distribuição por Status")
        status_distribution = report_data.get("status_distribution", [])

        ws_status["A1"] = "Distribuição de Conversas por Status"
        ws_status["A1"].font = Font(bold=True, size=14)  # type: ignore[index]

        status_headers = ["Status", "Quantidade", "Percentual (%)"]
        for col, header in enumerate(status_headers, start=1):
            cell = ws_status.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, status in enumerate(status_distribution, start=4):
            ws_status.cell(row=row_idx, column=1, value=status.get("status", "N/A"))
            ws_status.cell(row=row_idx, column=2, value=status.get("count", 0))
            ws_status.cell(row=row_idx, column=3, value=status.get("percentage", 0))
            ws_status.cell(row=row_idx, column=1).fill = data_fill  # type: ignore[index]
            ws_status.cell(row=row_idx, column=2).fill = data_fill  # type: ignore[index]
            ws_status.cell(row=row_idx, column=3).fill = data_fill  # type: ignore[index]

        ws_status.column_dimensions["A"].width = 25  # type: ignore[index]
        ws_status.column_dimensions["B"].width = 15  # type: ignore[index]
        ws_status.column_dimensions["C"].width = 15  # type: ignore[index]

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        logger.info("[SUCCESS] Excel gerado com sucesso (%s bytes)", len(excel_bytes))
        return excel_bytes
